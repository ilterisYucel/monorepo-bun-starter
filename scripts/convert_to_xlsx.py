#!/usr/bin/env python3
"""Convert register-ui-mapping.json to formatted .xlsx
Usage: python3 scripts/convert_to_xlsx.py [input.json] [output.xlsx]
"""

import json
import sys
import os

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

input_path = sys.argv[1] if len(sys.argv) > 1 else os.path.join(ROOT, 'docs', 'mappings', 'register-ui-mapping.json')
output_path = sys.argv[2] if len(sys.argv) > 2 else os.path.join(ROOT, 'docs', 'mappings', 'register-ui-mapping.xlsx')

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")
    print("Falling back to CSV output...")
    with open(input_path, encoding='utf-8') as f:
        rows = json.load(f)
    csv_path = os.path.join(ROOT, 'docs', 'mappings', 'register-ui-mapping.csv')
    import csv
    if rows:
        with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
            w = csv.DictWriter(f, fieldnames=rows[0].keys())
            w.writeheader()
            w.writerows(rows)
        print(f"Generated {len(rows)} rows → {csv_path} (CSV, opens in Excel)")
    sys.exit(0)

with open(input_path, encoding='utf-8') as f:
    rows = json.load(f)

if not rows:
    print("No rows to convert")
    sys.exit(1)

# ─── COLUMN DEFINITIONS ──────────────────────────────────────────────────────

COLUMNS = [
    ("Cihaz",         "device",         12),
    ("Cihaz ID",      "deviceId",       12),
    ("Cihaz Tipi",    "deviceType",     10),
    ("Register (Dec)","registerAddress", 14),
    ("Register (Hex)","registerHex",     12),
    ("Değişken Adı",  "variableName",   25),
    ("Data Tag",      "dataTag",        25),
    ("Harita Adı",    "mapName",        35),
    ("Kaynak",        "source",         20),
    ("Tip",           "mapType",         8),
    ("Birim",         "mapUnit",         6),
    ("Açıklama",      "mapDescription", 40),
    ("Telemetri Adı", "telemetryName",  25),
    ("Etiketler",     "tags",           30),
    ("Toplama",       "aggregation",    10),
    ("Kapsam",        "scope",          10),
    ("Raf/Birim",     "rackLabel",      15),
    ("Nerede",        "whereTo",        15),
    ("UI Bileşeni",   "uiComponent",    35),
    ("UI Sekme",      "uiTab",          16),
    ("UI Alan",       "uiField",        20),
    ("UI Etiket",     "uiFieldLabel",   30),
]

# ─── STYLES ───────────────────────────────────────────────────────────────────

wb = Workbook()
ws = wb.active
ws.title = "Register-UI Eşleştirme"

# Header style
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# Alternating row fills
even_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# Scope colors
scope_fills = {
    "system": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    "per-rack": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
    "per-unit": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
    "per-room": PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"),
}

# ─── WRITE HEADERS ───────────────────────────────────────────────────────────

for col_idx, (header, _, width) in enumerate(COLUMNS, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[1].height = 30

# ─── WRITE DATA ──────────────────────────────────────────────────────────────

for row_idx, row_data in enumerate(rows, 2):
    is_even = row_idx % 2 == 0
    scope = row_data.get('scope', '')

    for col_idx, (_, key, _) in enumerate(COLUMNS, 1):
        value = row_data.get(key, '')
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(name="Calibri", size=10)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Apply scope color to kapsam column
        if key == 'scope' and value in scope_fills:
            cell.fill = scope_fills[value]
            cell.font = Font(name="Calibri", size=10, bold=True)

        # Apply alternating row fill to non-scope cells
        if key != 'scope':
            cell.fill = even_fill if is_even else odd_fill

# ─── FREEZE HEADER & AUTO-FILTER ─────────────────────────────────────────────

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"

# ─── SAVE ─────────────────────────────────────────────────────────────────────

wb.save(output_path)
print(f"Generated {len(rows)} rows → {output_path}")
